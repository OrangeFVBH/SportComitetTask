from flask import Flask, jsonify, request
import openpyxl

app = Flask(__name__)
DATA_DB = []


def load_data_from_xlsx(path="C:\\Users\\Student\\Downloads\\sports_sections.xlsx"):
    try:
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]

        DATA_DB.clear()
        for row in ws.iter_rows(min_row=2, values_only=True):
            record = dict(zip(headers, row))
            if record.get("sports"):
                # Превращаем "Бокс; Футбол" в ["Бокс", "Футбол"]
                record["sports"] = [s.strip() for s in str(record["sports"]).split(";")]
            else:
                record["sports"] = []
            DATA_DB.append(record)
        print(f"Загружено записей: {len(DATA_DB)}")
    except Exception as e:
        print(f"Ошибка загрузки Excel: {e}")


load_data_from_xlsx()


@app.route('/api/sports', methods=['GET'])
def get_sports():
    all_sports = set()
    for section in DATA_DB:
        for sport in section['sports']:
            all_sports.add(sport.capitalize())
    return jsonify(sorted(list(all_sports)))


@app.route('/api/search', methods=['GET'])
def search_sections():
    # Очищаем входные данные от пробелов
    city_query = request.args.get('city', '').lower().replace(" ", "")
    sport_query = request.args.get('sport', '').lower().strip()

    results = []
    for section in DATA_DB:
        # Проверка города (удаляем пробелы из БД для сравнения)
        db_city = str(section.get('city', '')).lower().replace(" ", "")

        if city_query and city_query != db_city:
            continue

        # Поиск по видам спорта
        if any(sport_query in s.lower() for s in section['sports']):
            results.append(section)

    return jsonify(results)


if __name__ == '__main__':
    # host='0.0.0.0' позволяет подключаться с эмулятора (через 10.0.2.2) [cite: 21]
    app.run(host='0.0.0.0', port=5000, debug=True)